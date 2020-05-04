#!/usr/bin/python3

#updateCrypto.py - a utility for updating an .xlsx file which contains
#a portfolio of crypto currencies.
#
#By Colin Trierweiler.  Released under the GNU GPL
#
#This is a proof-of-concept, version 0.1, so to speak, but it works!
#
#It works by scraping all the h4 HTML tags off of Coinbase's price page
#and using certain indices to match certain currencies.  These indicies
#start at 3 for BTC, and then go up by 7 from there.  For the time being
#it will only work with the included CryptoPortfolio.xlsx spreadsheet.
#
#Future improvements will include:
# * Turning the scrape part into a function
# * Allowing for more customization of the contents of the portfolio
# * Turning it into a command line argument with an optional filename

#Set up

import openpyxl
import requests
import bs4

wb = openpyxl.load_workbook('CryptoPortfolio.xlsx')
sheet = wb['Sheet1']

res = requests.get('https://www.coinbase.com/price')
res.raise_for_status()
coinbaseSoup = bs4.BeautifulSoup(res.text, features='html5lib')

#Import all h4 tags (there has to be a better way!)

h4List = coinbaseSoup.select('h4')

#Scrape

newBTC = h4List[3].text[1:]
newETH = h4List[10].text[1:]
newXRP = h4List[17].text[1:]

if ',' in newBTC:
    newBTC = newBTC.replace(',', '')

#Update

PriceUpdates = {'Bitcoin': float(newBTC),
                 'Ether': float(newETH),
                 'Ripple': float(newXRP)}

for rowNum in range(2, sheet.max_row):
    currencyName = sheet.cell(row = rowNum, column = 1).value
    if currencyName in PriceUpdates:
        sheet.cell(row = rowNum, column = 3).value = PriceUpdates[currencyName]

wb.save('newCryptoPortfolio.xlsx')
